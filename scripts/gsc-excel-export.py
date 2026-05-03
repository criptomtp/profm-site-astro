#!/usr/bin/env python3
"""gsc-excel-export.py — Detailed GSC analytics → Excel workbook.

Generates a multi-sheet .xlsx with:
  - README (data limitations, methodology)
  - Summary (weekly site totals + WoW deltas)
  - All Queries (every query × weekly position + impression columns,
    plus auto trend classification)
  - Top Movers (biggest risers/fallers by impression delta)
  - Pages (every URL × weekly impressions/clicks/position)
  - Countries (geographic breakdown by week)
  - Devices (mobile/desktop/tablet split by week)

Time periods: monthly columns Jan-May 2026 (Jan/Feb/Mar will be empty
because GSC property was verified ~Apr 1 2026 — there is no pre-Apr data
in GSC for this site). PLUS weekly Apr W1-W4 + May W1 partial for actual
trend visibility.
"""

import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

ROOT = Path("/Users/nikolaj/My vibecode aplications/profm-site-astro")
sys.path.insert(0, str(ROOT / "scripts"))

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

SCOPES = [
    "https://www.googleapis.com/auth/webmasters",
    "https://www.googleapis.com/auth/indexing",
]
TOKEN_FILE = ROOT / "scripts" / "gsc_token.json"
SITE = "sc-domain:fulfillmentmtp.com.ua"
OUT_DIR = ROOT / "docs" / "gsc"
TODAY = date.today()

# Time periods we'll show as columns. Monthly Jan-May (Jan/Feb/Mar empty
# because GSC was added Apr) + weekly Apr W1-W4 + May W1 partial.
PERIODS_MONTHLY = [
    ("Jan 2026", "2026-01-01", "2026-01-31"),
    ("Feb 2026", "2026-02-01", "2026-02-28"),
    ("Mar 2026", "2026-03-01", "2026-03-31"),
    ("Apr 2026", "2026-04-01", "2026-04-30"),
    ("May 2026", "2026-05-01", "2026-05-02"),  # partial — GSC has 3-day lag
]

PERIODS_WEEKLY = [
    ("Apr W1 (1-7)",   "2026-04-01", "2026-04-07"),
    ("Apr W2 (8-14)",  "2026-04-08", "2026-04-14"),
    ("Apr W3 (15-21)", "2026-04-15", "2026-04-21"),
    ("Apr W4 (22-28)", "2026-04-22", "2026-04-28"),
    ("Apr-May W5",     "2026-04-29", "2026-05-02"),
]

# Style helpers
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
NA_FILL  = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def get_service():
    creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("searchconsole", "v1", credentials=creds)


def query(service, start, end, dimensions, row_limit=25000):
    body = {"startDate": start, "endDate": end, "dimensions": dimensions, "rowLimit": row_limit}
    resp = service.searchanalytics().query(siteUrl=SITE, body=body).execute()
    return resp.get("rows", [])


def site_totals(service, start, end):
    body = {"startDate": start, "endDate": end, "dimensions": []}
    resp = service.searchanalytics().query(siteUrl=SITE, body=body).execute()
    rows = resp.get("rows", [])
    if not rows:
        return 0, 0, 0, 0
    r = rows[0]
    imp = r.get("impressions", 0)
    clk = r.get("clicks", 0)
    return imp, clk, (clk / imp if imp else 0), r.get("position", 0)


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def autosize(ws, max_width=60):
    """Autosize columns. Handles MergedCell which lacks .column_letter."""
    col_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            # Skip merged cells (only the top-left of a merge range has a value)
            if not hasattr(cell, "column_letter"):
                continue
            if cell.value is None:
                continue
            col_letter = cell.column_letter
            length = len(str(cell.value))
            if length > col_widths.get(col_letter, 0):
                col_widths[col_letter] = length
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = min(width + 2, max_width)


# ============== Build sheets ==============

def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    notes = [
        ("MTP Group — fulfillmentmtp.com.ua", True, 16),
        ("GSC Analytics Export", True, 14),
        (f"Generated: {TODAY.isoformat()}", False, 11),
        ("", False, 11),
        ("⚠️  CRITICAL DATA LIMITATION", True, 13),
        ("", False, 11),
        ("GSC property `sc-domain:fulfillmentmtp.com.ua` was verified", False, 11),
        ("around early April 2026 (during the CF Pages migration from", False, 11),
        ("Vercel/Tilda hosting). Google Search Console retains data only", False, 11),
        ("from the date a property is verified.", False, 11),
        ("", False, 11),
        ("Result: Jan / Feb / Mar 2026 columns will be EMPTY (no GSC data).", True, 11),
        ("This is not a bug — the data simply does not exist in GSC.", False, 11),
        ("", False, 11),
        ("To get pre-April data, use one of:", False, 11),
        ("  • Google Analytics 4 — if it was running on the old hosting", False, 11),
        ("  • Tilda Analytics archives", False, 11),
        ("  • Server referrer logs (if retained)", False, 11),
        ("  • Wayback Machine for SERP screenshots (manual)", False, 11),
        ("", False, 11),
        ("WHAT THIS WORKBOOK CONTAINS", True, 13),
        ("", False, 11),
        ("1. Summary — weekly site-level totals + WoW changes", False, 11),
        ("2. All Queries — every query found in GSC × monthly + weekly columns", False, 11),
        ("                  with avg position and impressions per period", False, 11),
        ("3. Top Movers — queries with biggest impression / position changes", False, 11),
        ("4. Pages — every URL × weekly impressions, clicks, position, CTR", False, 11),
        ("5. Countries — geographic breakdown per week", False, 11),
        ("6. Devices — mobile / desktop / tablet per week", False, 11),
        ("", False, 11),
        ("METHODOLOGY", True, 13),
        ("", False, 11),
        ("• Position = weighted average across all queries (impression-weighted)", False, 11),
        ("• Lower position number = better rank (pos 1 = top of SERP)", False, 11),
        ("• 'Δ Imp' = change in impressions vs previous period", False, 11),
        ("• 'Δ Pos' = change in position; negative = better", False, 11),
        ("• Empty cells in monthly columns = no data (property too new)", False, 11),
        ("• 'May 2026' is partial (May 1-2 only — GSC has 3-day reporting lag)", False, 11),
        ("• Trend column auto-classifies: 🟢 RISING, 🔴 FALLING, ➖ STABLE, 🆕 NEW", False, 11),
        ("", False, 11),
        ("Re-generate anytime: python3 scripts/gsc-excel-export.py", False, 11),
    ]
    for i, (text, bold, size) in enumerate(notes, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 80


def build_summary(wb, monthly_data, weekly_data):
    ws = wb.create_sheet("Summary")
    # Monthly section
    ws["A1"] = "Monthly site totals (Jan/Feb/Mar EMPTY — GSC property added Apr 2026)"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:F1")

    headers = ["Period", "Impressions", "Clicks", "CTR", "Avg Position", "Note"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    row = 4
    for label, _, _ in PERIODS_MONTHLY:
        imp, clk, ctr, pos = monthly_data[label]["totals"]
        ws.cell(row=row, column=1, value=label)
        if imp == 0:
            for c in range(2, 6):
                ws.cell(row=row, column=c, value="—")
                ws.cell(row=row, column=c).fill = NA_FILL
            ws.cell(row=row, column=6, value="No GSC data")
            ws.cell(row=row, column=6).fill = NA_FILL
        else:
            ws.cell(row=row, column=2, value=imp)
            ws.cell(row=row, column=3, value=clk)
            ws.cell(row=row, column=4, value=round(ctr, 4))
            ws.cell(row=row, column=4).number_format = "0.00%"
            ws.cell(row=row, column=5, value=round(pos, 1))
            note = ""
            if "May" in label:
                note = "Partial (May 1-2 only)"
            ws.cell(row=row, column=6, value=note)
        row += 1

    # Weekly section
    row += 2
    ws.cell(row=row, column=1, value="Weekly slices (where actual data lives)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=6)

    row += 2
    weekly_headers = ["Week", "Days", "Impressions", "Clicks", "CTR", "Avg Position"]
    for i, h in enumerate(weekly_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(weekly_headers))

    for label, start, end in PERIODS_WEEKLY:
        row += 1
        imp, clk, ctr, pos = weekly_data[label]["totals"]
        days = (date.fromisoformat(end) - date.fromisoformat(start)).days + 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=days)
        if imp == 0:
            for c in range(3, 7):
                ws.cell(row=row, column=c, value="—")
                ws.cell(row=row, column=c).fill = NA_FILL
        else:
            ws.cell(row=row, column=3, value=imp)
            ws.cell(row=row, column=4, value=clk)
            ws.cell(row=row, column=5, value=round(ctr, 4))
            ws.cell(row=row, column=5).number_format = "0.00%"
            ws.cell(row=row, column=6, value=round(pos, 1))

    autosize(ws, max_width=40)


def classify_trend(positions, impressions):
    """Trend across W1-W4 (all 7-day weeks). W5 (4-day partial) is excluded from
    classification because comparing partial vs full week always shows decline.

    Args:
      positions: list of weekly avg positions (None for missing)
      impressions: list of weekly impressions (None for missing)
    """
    # Use only full 7-day weeks (first 4 elements = W1-W4)
    if len(positions) >= 5:
        positions = positions[:4]
        impressions = impressions[:4]
    valid = [(p, i) for p, i in zip(positions, impressions) if p is not None and p > 0 and i and i > 0]
    if len(valid) == 0:
        return "—"
    if len(valid) == 1:
        return "🆕 NEW"
    first_p = valid[0][0]
    last_p = valid[-1][0]
    first_i = valid[0][1]
    last_i = valid[-1][1]
    pos_delta = last_p - first_p
    imp_delta_pct = (last_i - first_i) / first_i * 100 if first_i else 0
    # Both signals must align for a clear verdict; weight by sample size
    if first_i >= 5 and last_i >= 5:
        if pos_delta < -3 or imp_delta_pct > 50:
            return "🟢 RISING"
        if pos_delta > 3 or imp_delta_pct < -50:
            return "🔴 FALLING"
    elif imp_delta_pct > 100:
        return "🟢 RISING"
    elif imp_delta_pct < -50 and last_i >= 3:
        return "🔴 FALLING"
    return "➖ STABLE"


def build_all_queries(wb, monthly_data, weekly_data):
    ws = wb.create_sheet("All Queries")

    # Build master query list (union across all periods)
    all_queries = set()
    for label, _, _ in PERIODS_MONTHLY:
        for r in monthly_data[label].get("rows_q", []):
            all_queries.add(r["keys"][0])
    for label, _, _ in PERIODS_WEEKLY:
        for r in weekly_data[label].get("rows_q", []):
            all_queries.add(r["keys"][0])

    # Headers: Query, Trend, then for each period: Pos + Imp + Clk
    period_labels = [p[0] for p in PERIODS_MONTHLY] + [p[0] for p in PERIODS_WEEKLY]
    headers = ["Query", "Trend", "Total Imp (W1-W5)"]
    for label in period_labels:
        headers.extend([f"{label} Pos", f"{label} Imp", f"{label} Clk"])

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    # Build per-query maps
    period_qmaps = {}
    for label, _, _ in PERIODS_MONTHLY:
        period_qmaps[label] = {r["keys"][0]: r for r in monthly_data[label].get("rows_q", [])}
    for label, _, _ in PERIODS_WEEKLY:
        period_qmaps[label] = {r["keys"][0]: r for r in weekly_data[label].get("rows_q", [])}

    # Sort queries by total weekly impressions (descending)
    def query_total_imp(q):
        return sum(period_qmaps[label].get(q, {}).get("impressions", 0) for label, _, _ in PERIODS_WEEKLY)

    sorted_queries = sorted(all_queries, key=lambda q: -query_total_imp(q))

    row = 2
    for q in sorted_queries:
        ws.cell(row=row, column=1, value=q)
        # Collect positions and impressions for trend
        weekly_positions = []
        weekly_imps = []
        for label, _, _ in PERIODS_WEEKLY:
            r = period_qmaps[label].get(q)
            if r:
                weekly_positions.append(r.get("position", 0))
                weekly_imps.append(r.get("impressions", 0))
            else:
                weekly_positions.append(None)
                weekly_imps.append(None)
        trend = classify_trend(weekly_positions, weekly_imps)
        ws.cell(row=row, column=2, value=trend)
        ws.cell(row=row, column=3, value=query_total_imp(q))

        col = 4
        for label in period_labels:
            r = period_qmaps[label].get(q)
            if r:
                ws.cell(row=row, column=col, value=round(r.get("position", 0), 1))
                ws.cell(row=row, column=col + 1, value=r.get("impressions", 0))
                ws.cell(row=row, column=col + 2, value=r.get("clicks", 0))
            else:
                # Mark visually as N/A for empty months
                if "Jan" in label or "Feb" in label or "Mar" in label:
                    for c_off in range(3):
                        ws.cell(row=row, column=col + c_off).fill = NA_FILL
            col += 3
        row += 1

    # Freeze header row + first 2 columns
    ws.freeze_panes = "D2"
    autosize(ws, max_width=55)
    # But cap query column width
    ws.column_dimensions["A"].width = 50


def build_top_movers(wb, weekly_data):
    ws = wb.create_sheet("Top Movers")
    w1_label = next((w[0] for w in PERIODS_WEEKLY if weekly_data[w[0]]["totals"][0] > 0), PERIODS_WEEKLY[0][0])
    last_full_label = "Apr W4 (22-28)"

    ws["A1"] = f"Comparing {w1_label} → {last_full_label}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:H1")

    qmap_first = {r["keys"][0]: r for r in weekly_data[w1_label].get("rows_q", [])}
    qmap_last = {r["keys"][0]: r for r in weekly_data[last_full_label].get("rows_q", [])}

    # Movements (queries in BOTH)
    movements = []
    for q in set(qmap_first) & set(qmap_last):
        f = qmap_first[q]
        l = qmap_last[q]
        f_imp = f.get("impressions", 0)
        l_imp = l.get("impressions", 0)
        f_pos = f.get("position", 0)
        l_pos = l.get("position", 0)
        if f_imp >= 3 or l_imp >= 3:
            movements.append((q, f_imp, l_imp, f_pos, l_pos, l_imp - f_imp, l_pos - f_pos))

    # NEW queries
    new_queries = []
    for q in set(qmap_last) - set(qmap_first):
        l = qmap_last[q]
        if l.get("impressions", 0) >= 3:
            new_queries.append((q, l.get("impressions", 0), l.get("position", 0)))
    new_queries.sort(key=lambda x: -x[1])

    # LOST queries
    lost_queries = []
    for q in set(qmap_first) - set(qmap_last):
        f = qmap_first[q]
        if f.get("impressions", 0) >= 3:
            lost_queries.append((q, f.get("impressions", 0), f.get("position", 0)))
    lost_queries.sort(key=lambda x: -x[1])

    row = 3
    # 🟢 Risers
    ws.cell(row=row, column=1, value="🟢 TOP RISERS (gained impressions)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="00B050")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=8)
    row += 1
    headers = ["Query", f"{w1_label} Imp", f"{last_full_label} Imp", "Δ Imp", f"{w1_label} Pos", f"{last_full_label} Pos", "Δ Pos", "Note"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    risers = sorted(movements, key=lambda m: -m[5])[:30]
    for q, fi, li, fp, lp, di, dp in risers:
        ws.cell(row=row, column=1, value=q)
        ws.cell(row=row, column=2, value=fi)
        ws.cell(row=row, column=3, value=li)
        ws.cell(row=row, column=4, value=di)
        ws.cell(row=row, column=5, value=round(fp, 1))
        ws.cell(row=row, column=6, value=round(lp, 1))
        ws.cell(row=row, column=7, value=round(dp, 1))
        ws.cell(row=row, column=8, value="✅ Better pos" if dp < -1 else ("⚠️ Worse pos" if dp > 1 else ""))
        row += 1

    # 🔴 Fallers
    row += 2
    ws.cell(row=row, column=1, value="🔴 TOP FALLERS (lost impressions)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="C00000")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=8)
    row += 1
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    fallers = sorted(movements, key=lambda m: m[5])[:30]
    for q, fi, li, fp, lp, di, dp in fallers:
        ws.cell(row=row, column=1, value=q)
        ws.cell(row=row, column=2, value=fi)
        ws.cell(row=row, column=3, value=li)
        ws.cell(row=row, column=4, value=di)
        ws.cell(row=row, column=5, value=round(fp, 1))
        ws.cell(row=row, column=6, value=round(lp, 1))
        ws.cell(row=row, column=7, value=round(dp, 1))
        ws.cell(row=row, column=8, value="✅ Better pos" if dp < -1 else ("⚠️ Worse pos" if dp > 1 else ""))
        row += 1

    # 🆕 New queries
    row += 2
    ws.cell(row=row, column=1, value=f"🆕 NEW queries (appeared in {last_full_label}, not in {w1_label})")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="0070C0")
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=8)
    row += 1
    new_headers = ["Query", "Imp", "Pos", "Clk"]
    for i, h in enumerate(new_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(new_headers))
    row += 1
    for q, imp, pos in new_queries[:30]:
        ws.cell(row=row, column=1, value=q)
        ws.cell(row=row, column=2, value=imp)
        ws.cell(row=row, column=3, value=round(pos, 1))
        ws.cell(row=row, column=4, value=qmap_last[q].get("clicks", 0))
        row += 1

    # 👻 Lost queries
    if lost_queries:
        row += 2
        ws.cell(row=row, column=1, value=f"👻 LOST queries (in {w1_label}, gone by {last_full_label})")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="7030A0")
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=8)
        row += 1
        for i, h in enumerate(new_headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(new_headers))
        row += 1
        for q, imp, pos in lost_queries[:30]:
            ws.cell(row=row, column=1, value=q)
            ws.cell(row=row, column=2, value=imp)
            ws.cell(row=row, column=3, value=round(pos, 1))
            ws.cell(row=row, column=4, value=qmap_first[q].get("clicks", 0))
            row += 1

    autosize(ws, max_width=55)
    ws.column_dimensions["A"].width = 50


def build_pages(wb, weekly_data):
    ws = wb.create_sheet("Pages")
    all_pages = set()
    for label, _, _ in PERIODS_WEEKLY:
        for r in weekly_data[label].get("rows_p", []):
            all_pages.add(r["keys"][0])

    period_pmaps = {label: {r["keys"][0]: r for r in weekly_data[label].get("rows_p", [])} for label, _, _ in PERIODS_WEEKLY}

    headers = ["Page", "Trend", "Total Imp"]
    for label, _, _ in PERIODS_WEEKLY:
        headers.extend([f"{label} Imp", f"{label} Clk", f"{label} Pos"])
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    def page_total_imp(p):
        return sum(period_pmaps[label].get(p, {}).get("impressions", 0) for label, _, _ in PERIODS_WEEKLY)

    sorted_pages = sorted(all_pages, key=lambda p: -page_total_imp(p))

    row = 2
    for p in sorted_pages:
        ws.cell(row=row, column=1, value=p.replace("https://www.fulfillmentmtp.com.ua", "") or "/")
        weekly_pos = []
        weekly_imp = []
        for label, _, _ in PERIODS_WEEKLY:
            r = period_pmaps[label].get(p)
            if r:
                weekly_pos.append(r.get("position", 0))
                weekly_imp.append(r.get("impressions", 0))
            else:
                weekly_pos.append(None)
                weekly_imp.append(None)
        ws.cell(row=row, column=2, value=classify_trend(weekly_pos, weekly_imp))
        ws.cell(row=row, column=3, value=page_total_imp(p))

        col = 4
        for label, _, _ in PERIODS_WEEKLY:
            r = period_pmaps[label].get(p)
            if r:
                ws.cell(row=row, column=col, value=r.get("impressions", 0))
                ws.cell(row=row, column=col + 1, value=r.get("clicks", 0))
                ws.cell(row=row, column=col + 2, value=round(r.get("position", 0), 1))
            col += 3
        row += 1

    ws.freeze_panes = "D2"
    autosize(ws, max_width=70)
    ws.column_dimensions["A"].width = 60


def build_countries_devices(wb, weekly_data):
    # Countries
    ws = wb.create_sheet("Countries")
    all_c = set()
    for label, _, _ in PERIODS_WEEKLY:
        for r in weekly_data[label].get("rows_c", []):
            all_c.add(r["keys"][0])
    period_cmaps = {label: {r["keys"][0]: r for r in weekly_data[label].get("rows_c", [])} for label, _, _ in PERIODS_WEEKLY}

    headers = ["Country", "Total Imp"]
    for label, _, _ in PERIODS_WEEKLY:
        headers.extend([f"{label} Imp", f"{label} Clk", f"{label} CTR"])
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))

    def c_total(c):
        return sum(period_cmaps[label].get(c, {}).get("impressions", 0) for label, _, _ in PERIODS_WEEKLY)

    row = 2
    for c in sorted(all_c, key=lambda x: -c_total(x)):
        ws.cell(row=row, column=1, value=c.upper())
        ws.cell(row=row, column=2, value=c_total(c))
        col = 3
        for label, _, _ in PERIODS_WEEKLY:
            r = period_cmaps[label].get(c)
            if r:
                imp = r.get("impressions", 0)
                clk = r.get("clicks", 0)
                ctr = clk / imp if imp else 0
                ws.cell(row=row, column=col, value=imp)
                ws.cell(row=row, column=col + 1, value=clk)
                ws.cell(row=row, column=col + 2, value=round(ctr, 4))
                ws.cell(row=row, column=col + 2).number_format = "0.00%"
            col += 3
        row += 1
    autosize(ws, max_width=20)

    # Devices
    ws = wb.create_sheet("Devices")
    all_d = ["MOBILE", "DESKTOP", "TABLET"]
    period_dmaps = {label: {r["keys"][0]: r for r in weekly_data[label].get("rows_d", [])} for label, _, _ in PERIODS_WEEKLY}

    headers = ["Device", "Total Imp"]
    for label, _, _ in PERIODS_WEEKLY:
        headers.extend([f"{label} Imp", f"{label} Clk", f"{label} CTR", f"{label} Pos"])
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))

    def d_total(d):
        return sum(period_dmaps[label].get(d, {}).get("impressions", 0) for label, _, _ in PERIODS_WEEKLY)

    row = 2
    for d in all_d:
        ws.cell(row=row, column=1, value=d)
        ws.cell(row=row, column=2, value=d_total(d))
        col = 3
        for label, _, _ in PERIODS_WEEKLY:
            r = period_dmaps[label].get(d)
            if r:
                imp = r.get("impressions", 0)
                clk = r.get("clicks", 0)
                ctr = clk / imp if imp else 0
                pos = r.get("position", 0)
                ws.cell(row=row, column=col, value=imp)
                ws.cell(row=row, column=col + 1, value=clk)
                ws.cell(row=row, column=col + 2, value=round(ctr, 4))
                ws.cell(row=row, column=col + 2).number_format = "0.00%"
                ws.cell(row=row, column=col + 3, value=round(pos, 1))
            col += 4
        row += 1
    autosize(ws, max_width=20)


# ============== Main ==============

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    service = get_service()

    print(f"Pulling monthly data ({len(PERIODS_MONTHLY)} months)...")
    monthly_data = {}
    for label, start, end in PERIODS_MONTHLY:
        print(f"  {label}: {start} → {end}")
        monthly_data[label] = {
            "totals": site_totals(service, start, end),
            "rows_q": query(service, start, end, ["query"], 25000),
            "rows_p": query(service, start, end, ["page"], 5000),
        }

    print(f"\nPulling weekly data ({len(PERIODS_WEEKLY)} weeks)...")
    weekly_data = {}
    for label, start, end in PERIODS_WEEKLY:
        print(f"  {label}: {start} → {end}")
        weekly_data[label] = {
            "totals": site_totals(service, start, end),
            "rows_q": query(service, start, end, ["query"], 25000),
            "rows_p": query(service, start, end, ["page"], 5000),
            "rows_c": query(service, start, end, ["country"], 250),
            "rows_d": query(service, start, end, ["device"], 10),
        }

    print("\nBuilding workbook...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_readme(wb)
    print("  ✅ README sheet")
    build_summary(wb, monthly_data, weekly_data)
    print("  ✅ Summary sheet")
    build_all_queries(wb, monthly_data, weekly_data)
    print("  ✅ All Queries sheet")
    build_top_movers(wb, weekly_data)
    print("  ✅ Top Movers sheet")
    build_pages(wb, weekly_data)
    print("  ✅ Pages sheet")
    build_countries_devices(wb, weekly_data)
    print("  ✅ Countries + Devices sheets")

    out_file = OUT_DIR / f"gsc-analytics-{TODAY.isoformat()}.xlsx"
    wb.save(out_file)
    print(f"\n✅ Excel saved: {out_file.relative_to(ROOT)}")
    print(f"   File size: {out_file.stat().st_size:,} bytes")
    return out_file


if __name__ == "__main__":
    main()
