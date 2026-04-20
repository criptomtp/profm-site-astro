#!/usr/bin/env python3
"""Generate Excel workbook from semantic core roadmap docs + GSC data.
Output: docs/SEMANTIC_CORE_ROADMAP.xlsx (8 sheets).
"""
import os, csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.comments import Comment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'docs', 'SEMANTIC_CORE_ROADMAP.xlsx')
PROGRESS_FILE = os.path.join(ROOT, 'docs', 'semantic-core-progress.json')

# Load progress log — maps task name → {status, date, commit, notes}
progress_map = {}
if os.path.exists(PROGRESS_FILE):
    with open(PROGRESS_FILE, encoding='utf-8') as f:
        _p = json.load(f)
        for c in _p.get('completions', []):
            progress_map[c['task']] = c

wb = Workbook()

# ===== STYLES =====
HEADER_FILL = PatternFill('solid', fgColor='000000')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
P0_FILL = PatternFill('solid', fgColor='FFCCCC')  # red
P1_FILL = PatternFill('solid', fgColor='FFE4B5')  # orange
P2_FILL = PatternFill('solid', fgColor='FFF9C4')  # yellow
DONE_FILL = PatternFill('solid', fgColor='C8E6C9')  # green
GRAY_FILL = PatternFill('solid', fgColor='EEEEEE')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(vertical='top', wrap_text=True)
THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell='A2'):
    ws.freeze_panes = cell

# ===== SHEET 1: README =====
ws = wb.active
ws.title = '📖 README'
ws['A1'] = 'SEMANTIC CORE ROADMAP — MTP Group (fulfillmentmtp.com.ua)'
ws['A1'].font = Font(bold=True, size=16, color='E63329')
ws.merge_cells('A1:E1')
ws['A2'] = 'Дата: 2026-04-20 | Джерело: GSC 90d + Competitor research + Actual site inventory'
ws['A2'].font = Font(italic=True, size=10, color='666666')
ws.merge_cells('A2:E2')

readme = [
    [''],
    ['📑 ЗМІСТ ФАЙЛУ'],
    [''],
    ['Sheet', 'Що всередині', 'Коли дивитись'],
    ['1. 📖 README', 'Опис книги + легенда', 'Зараз'],
    ['2. 📊 Dashboard', 'Зведена статистика за 90d GSC + КPI цілі', 'Старт кожного тижня'],
    ['3. ✅ Current inventory', 'Всі 170 deployed pages з GSC метриками', 'Коли шукаєш "що вже є"'],
    ['4. 🎯 Striking distance', '9 opportunities pos 4-20 з GSC — швидкі перемоги', 'Коли вирішуєш що оптимізувати'],
    ['5. ⚠️ Gap analysis', 'Сторінки яких бракує (з GSC+competitor)', 'Коли плануєш створення нових'],
    ['6. 🗓️ Roadmap Phases', 'План 6 місяців по фазах P0→P2', 'Coли плануєш спринт'],
    ['7. 🔄 v1 vs v2', 'Що підтверджено / спростовано / нове', 'Коли питаєшся чому план змінився'],
    ['8. 🏢 Competitors', 'Аналіз 20+ конкурентів + semantic gaps', 'Коли вивчаєш ринок'],
    ['9. 🔍 Top queries', 'Всі 334 GSC queries за 90d', 'Коли шукаєш специфічний keyword'],
    [''],
    ['🎨 ЛЕГЕНДА КОЛЬОРІВ'],
    [''],
    ['P0 — критично', '', 'Треба робити СЬОГОДНІ/ЦЕЙ ТИЖДЕНЬ'],
    ['P1 — важливо', '', 'Треба зробити цього місяця'],
    ['P2 — бажано', '', 'Backlog, коли буде час'],
    ['✅ Зроблено', '', 'Deployed in production'],
    [''],
    ['📂 ПОВ\'ЯЗАНІ ФАЙЛИ (детальні дані)'],
    [''],
    ['docs/SEMANTIC_CORE_ROADMAP.md', 'Основний документ (читати перед цим Excel)'],
    ['docs/SEMANTIC_CORE_GSC_AUDIT.md', 'Детальний GSC аудит з форекастом'],
    ['docs/MTP_SEMANTIC_CORE_FULL.md', 'План v1 від 2026-03-27 (для порівняння)'],
    ['docs/gsc/competitor-gap-research.md', 'Аналіз 20+ конкурентів (сирий)'],
    ['docs/gsc/full-pages.csv / .json', 'Сирі GSC дані — 178 сторінок'],
    ['docs/gsc/full-queries.csv / .json', 'Сирі GSC дані — 334 queries'],
    ['docs/gsc/opportunities.csv', 'Сирі striking-distance дані'],
]

start_row = 3
for r in readme:
    ws.append(r)

# Style TOC header
for cell in ws[6]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER

# Style legend rows
for row_idx in [17, 18, 19, 20]:
    for cell in ws[row_idx]:
        cell.alignment = WRAP

ws.row_dimensions[17].height = 20
ws['A17'].fill = P0_FILL
ws['A18'].fill = P1_FILL
ws['A19'].fill = P2_FILL
ws['A20'].fill = DONE_FILL

set_widths(ws, [30, 45, 55, 15, 15])

# ===== SHEET 2: DASHBOARD =====
ws = wb.create_sheet('📊 Dashboard')
ws['A1'] = 'DASHBOARD — 90 Days (GSC baseline 2026-01-20 → 2026-04-20)'
ws['A1'].font = Font(bold=True, size=14, color='E63329')
ws.merge_cells('A1:D1')

dashboard = [
    [''],
    ['МЕТРИКА', 'ПОТОЧНЕ', 'ЦІЛЬ 6 МІС', 'ПРИРІСТ'],
    ['Total clicks (90d)', 29, '4500-9000', '150-310x'],
    ['Total impressions (90d)', 4329, '300000+', '70x'],
    ['CTR середній', '0.67%', '1.5-2.5%', '+2.2-3.7x'],
    ['Середня позиція', 15, 'Топ-5', '3x улучшення'],
    ['Total pages deployed', 170, '200-220', '+30-50'],
    ['Pages з органічним трафіком', '~10', '80-100', '+10x'],
    [''],
    ['РОЗПОДІЛ ПО МОВАХ (90d GSC)'],
    ['Мова', 'Pages', 'Clicks', 'Impressions', 'CTR'],
    ['UA', 23, 16, 1547, '1.03%'],
    ['RU', 19, 2, 590, '0.34%'],
    ['EN', 55, 1, 591, '0.17%'],
    ['Root (neutral)', 73, 10, 1601, '0.62%'],
    [''],
    ['ТОП-5 ПРОБЛЕМ'],
    ['#', 'Проблема', 'Impact', 'Action'],
    [1, 'EN 55 pages → 1 клік (семантика не та)', 'Критичний — ресурси витрачені даремно', 'Заморозити + 1 deep pillar EN'],
    [2, '"фулфілмент" pos 16 (2-га сторінка)', 'Головний keyword галузі втрачений', '/ua/shcho-take-fulfilment/ pillar'],
    [3, '"mtp" brand query 0 CTR на pos 3.8', 'Власний бренд не конвертує', 'Organization schema + GBP claim'],
    [4, '"фулфилмент цены" 0 CTR на pos 2.5', 'Buyer intent + top pos = втрата', 'Meta з ціною "від 9 грн"'],
    [5, 'RU blog = 0 posts (30% ринку)', 'Цілий сегмент відрізаний', 'Phase 2: 3 RU posts'],
    [''],
    ['ТОП-5 ШВИДКИХ ПЕРЕМОГ'],
    ['#', 'Дія', 'Effort', 'Очікуваний ефект'],
    [1, 'Home title + meta з ціною', '30 хв', '+40-60 clicks/mo'],
    [2, 'Organization schema для brand', '1 год', '+20-40 clicks/mo'],
    [3, '/ua/blog/scho-take-artykul-sku/ (pos 5.5)', '1 день', '+30-50 clicks/mo'],
    [4, '/ua/blog/scho-take-sla-v-logistici/ (pos 8.2)', '1 день', '+25-40 clicks/mo'],
    [5, '/ua/shcho-take-fulfilment/ pillar', '3-5 днів', '+100-200 clicks/mo через 2-3 міс'],
]

for r in dashboard:
    ws.append(r)

# Style subheaders
for row_idx, row_type in [(3, 'h'), (11, 'h'), (12, 'h'), (18, 'h'), (19, 'h'), (26, 'h'), (27, 'h')]:
    for cell in ws[row_idx]:
        if cell.value:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

# KPI cells highlight
for row_idx in range(4, 10):
    ws[f'B{row_idx}'].fill = P0_FILL
    ws[f'C{row_idx}'].fill = DONE_FILL

set_widths(ws, [48, 20, 25, 50])

# ===== SHEET 3: CURRENT INVENTORY =====
ws = wb.create_sheet('✅ Current inventory')

# Load full-pages.csv
pages_file = os.path.join(ROOT, 'docs', 'gsc', 'full-pages.csv')
pages = []
with open(pages_file, encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for r in reader:
        pages.append(r)

ws.append(['URL', 'Lang', 'Type', 'Clicks 90d', 'Impressions 90d', 'CTR %', 'Avg pos', 'Top queries', 'Status', 'Action needed'])

def classify_url(url):
    u = url.lower()
    if '/ua/' in u: lang = 'UA'
    elif '/ru/' in u: lang = 'RU'
    elif '/en/' in u: lang = 'EN'
    else: lang = 'ROOT'

    if '/blog/' in u: typ = 'blog'
    elif 'calculator' in u: typ = 'calculator'
    elif 'tsiny' in u or 'prices' in u or 'tsenu' in u: typ = 'pricing'
    elif 'faq' in u: typ = 'FAQ'
    elif 'about' in u: typ = 'about'
    elif 'fulfilment-dlya' in u or 'fulfillment-for' in u: typ = 'vertical'
    elif url.endswith('/ua/') or url.endswith('/ru/') or url.endswith('/en/') or url.endswith('/'): typ = 'home'
    elif 'thanks' in u: typ = 'thanks'
    else: typ = 'landing'
    return lang, typ

def status_action(row, lang, typ):
    try:
        clk = int(row['clicks_90d'])
        imp = int(row['impressions_90d'])
        pos = float(row['avg_position']) if row['avg_position'] else 99
    except:
        clk, imp, pos = 0, 0, 99

    if clk >= 3:
        return 'WORKING', 'Оптимізувати CTR' if pos < 5 else 'Push до топ-3'
    if imp >= 100 and pos >= 4 and pos <= 20:
        return 'STRIKING DISTANCE', f'Push pos {pos:.1f} → топ-3'
    if imp >= 50 and pos > 20:
        return 'VISIBLE BUT FAR', 'Content update + backlinks'
    if imp < 10:
        if typ == 'blog' and lang == 'EN':
            return 'NO TRAFFIC EN', 'Candidate to archive/301'
        return 'NO TRAFFIC', 'Перевірити intent'
    return 'LOW TRAFFIC', 'Moніторити'

for r in pages:
    lang, typ = classify_url(r['url'])
    status, action = status_action(r, lang, typ)
    ws.append([
        r['url'], lang, typ,
        int(r['clicks_90d']) if r['clicks_90d'] else 0,
        int(r['impressions_90d']) if r['impressions_90d'] else 0,
        float(r['ctr_pct']) if r['ctr_pct'] else 0,
        float(r['avg_position']) if r['avg_position'] else 99,
        r.get('top_queries', '')[:200],
        status, action
    ])

style_header(ws)
freeze(ws, 'A2')
set_widths(ws, [65, 8, 14, 12, 14, 10, 10, 50, 22, 30])

# Status color-coding
for row_idx in range(2, ws.max_row + 1):
    status = ws[f'I{row_idx}'].value
    if status == 'WORKING':
        ws[f'I{row_idx}'].fill = DONE_FILL
    elif status == 'STRIKING DISTANCE':
        ws[f'I{row_idx}'].fill = P0_FILL
    elif status in ('NO TRAFFIC EN', 'VISIBLE BUT FAR'):
        ws[f'I{row_idx}'].fill = P1_FILL
    elif status == 'LOW TRAFFIC':
        ws[f'I{row_idx}'].fill = P2_FILL

# Sort by impressions desc — done via formula later in Excel; data already sorted from source

# Apply autofilter
ws.auto_filter.ref = f'A1:J{ws.max_row}'

# ===== SHEET 4: STRIKING DISTANCE =====
ws = wb.create_sheet('🎯 Striking distance')

opps_file = os.path.join(ROOT, 'docs', 'gsc', 'opportunities.csv')
opps = []
with open(opps_file, encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for r in reader:
        opps.append(r)

ws.append(['#', 'Query', 'Current page', 'Pos', 'Impressions 90d', 'Clicks 90d', 'CTR %', 'Score', 'Action', 'Owner page (new or existing)', 'Priority', 'Est. effort', 'Est. clicks/mo if top-3'])

# Map opps to actions based on my analysis
action_map = {
    'мтп': ('Organization schema + GBP claim для brand', '/ua/ + /', 'P0', '2h', '30-50'),
    'артикул': ('Створити повноцінний гайд', '/ua/blog/scho-take-artykul-sku/', 'P0', '1 день', '40-60'),
    'sla': ('Створити повноцінний гайд', '/ua/blog/scho-take-sla-v-logistici/', 'P0', '1 день', '30-50'),
    'sla це': ('Об\'єднати в /scho-take-sla-v-logistici/', '/ua/blog/scho-take-sla-v-logistici/', 'P0', 'в складі попередньої', '15-25'),
    'товарний бізнес': ('Переписати deep 2K слів', '/ua/blog/tovarnyi-biznes/', 'P1', '1 день', '20-35'),
    'товарный бизнес': ('RU версія, переписати deep', '/ru/blog/tovarnyi-biznes/', 'P1', '1 день', '15-25'),
    'фулфілмент': ('Pillar hub 3-4K слів', '/ua/shcho-take-fulfilment/', 'P0', '3-5 днів', '100-200'),
}

for i, r in enumerate(opps, 1):
    q = r['query']
    action, owner, prio, effort, est_clicks = action_map.get(q, ('Оптимізувати контент + internal links', r['page'], 'P1', '2-3h', '10-20'))
    ws.append([
        i, q, r['page'],
        float(r['avg_position']),
        int(r['impressions']),
        int(r['clicks']),
        float(r['ctr_pct']),
        float(r['opportunity_score']),
        action, owner, prio, effort, est_clicks
    ])

style_header(ws)
freeze(ws, 'A2')
set_widths(ws, [5, 25, 60, 8, 14, 10, 8, 10, 40, 50, 10, 15, 20])

for row_idx in range(2, ws.max_row + 1):
    prio = ws[f'K{row_idx}'].value
    if prio == 'P0':
        ws[f'K{row_idx}'].fill = P0_FILL
    elif prio == 'P1':
        ws[f'K{row_idx}'].fill = P1_FILL
    elif prio == 'P2':
        ws[f'K{row_idx}'].fill = P2_FILL

ws.auto_filter.ref = f'A1:M{ws.max_row}'

# ===== SHEET 5: GAP ANALYSIS =====
ws = wb.create_sheet('⚠️ Gap analysis')

gaps = [
    ['Category', 'Gap / Missing page', 'Query (monthly imp)', 'Priority', 'Phase', 'Suggested URL', 'Notes'],
    # Pillar
    ['PILLAR', '"Що таке фулфілмент" hub', '1000-2000/mo', 'P0', 'Phase 1', '/ua/shcho-take-fulfilment/', 'Cross-link у 22 вертикальні + калькулятор. 3-4K слів.'],
    # Striking
    ['QUICK WIN', '"Артикул це" гайд', '336 (GSC)', 'P0', 'Phase 1', '/ua/blog/scho-take-artykul-sku/', 'Pos 5.5 зараз — push в топ-3'],
    ['QUICK WIN', '"SLA в логістиці"', '365 (GSC)', 'P0', 'Phase 1', '/ua/blog/scho-take-sla-v-logistici/', 'Pos 8.2 зараз'],
    # Geography (NEW in v2)
    ['GEO', 'Дніпро фулфілмент', '72 (GSC) + est 300+', 'P0', 'Phase 1', '/ua/fulfilment-dnipro/', 'Gap НЕ був у v1 плані'],
    ['GEO', 'Харків фулфілмент', '38 (GSC) + est 200+', 'P1', 'Phase 2', '/ua/fulfilment-kharkiv/', 'Не було в v1'],
    ['GEO', 'Львів фулфілмент (повна)', '55 (GSC)', 'P1', 'Phase 2', '/ua/fulfilment-lviv/', 'v1 мав "чому Київ кращий" — помилка'],
    ['GEO', 'Одеса фулфілмент (повна)', '45 (GSC)', 'P1', 'Phase 2', '/ua/fulfilment-odesa/', 'v1 мав "чому Київ кращий" — помилка'],
    # RU blog
    ['RU BLOG', 'RU blog infrastructure', '~30% ринку UA', 'P1', 'Phase 2', '/ru/blog/', 'Зараз 0 posts, 0 трафіку'],
    ['RU BLOG', 'Что такое фулфилмент (RU)', 'est 500+', 'P1', 'Phase 2', '/ru/blog/chto-takoye-fulfilment/', 'RU дублікат UA pillar'],
    ['RU BLOG', 'Как открыть интернет-магазин', 'existing Tilda redirect', 'P1', 'Phase 2', '/ru/blog/kak-otkryt-internet-magazin/', 'Був на Tilda, зараз 301'],
    ['RU BLOG', 'SLA в логистике (RU)', 'existing Tilda', 'P1', 'Phase 2', '/ru/blog/sla-v-logistike/', 'Був на Tilda'],
    # Marketplace (from competitor research)
    ['MARKETPLACE', 'Розетка fulfillment', 'competitor gap', 'P1', 'Phase 2', '/ua/fulfilment-rozetka/', 'Competitor research виявив окрему вісь'],
    ['MARKETPLACE', 'Prom.ua fulfillment', 'competitor gap', 'P1', 'Phase 2', '/ua/fulfilment-prom/', 'Окрема landing, не в складі загального'],
    ['MARKETPLACE', 'OLX fulfillment', 'competitor gap', 'P1', 'Phase 2', '/ua/fulfilment-olx/', 'OLX доставка specifically'],
    # Pricing
    ['PRICING', 'Публічна прайс-таблиця', '"фулфилмент цены" 0 CTR pos 2.5', 'P1', 'Phase 3', '/ua/tarify/', 'Не калькулятор — буквально таблиця з цифрами'],
    # Consolidation (NEW strategy)
    ['CONSOL', '22 vertical pages → 8 deep hubs', 'existing 6 have 0 traffic', 'P1', 'Phase 3', '/ua/fulfilment-{beauty,fashion,electronics,kids,home,handmade,auto,food}/', 'Radical consolidation, проти v1'],
    # Competitor gaps
    ['CONTENT', '"Відповідальне зберігання" hub', '500-1000/mo', 'P1', 'Phase 3', '/ua/vidpovidalne-zberigannya/', 'Competitor research — MTP не має'],
    ['CONTENT', 'ТОП фулфілмент операторів України', 'listicle queries', 'P1', 'Phase 3', '/ua/blog/top-fulfilment-ukrainy/', 'v1 мав як 🟠, досі не зроблено'],
    # EN rebuild
    ['EN REBUILD', 'Deep EN pillar (1 не 18)', '55 existing → 1 click', 'P2', 'Phase 4', '/en/what-is-fulfillment-ukraine/', 'v2 радикально: 1 глибока > 18 тонких'],
    ['EN REBUILD', 'Amazon FBA Prep', 'approved Q2 prior session', 'P2', 'Phase 4', '/en/amazon-fba-prep-ukraine/', 'Узгоджено в попередній сесії'],
    # CTR fixes (not new pages — existing optimization)
    ['CTR FIX', 'Home title + meta з ціною', '"фулфилмент цены" 0 CTR', 'P0', 'Phase 0', '/ua/ + /ru/ + /en/', '30 хв роботи, швидкий CTR boost'],
    ['CTR FIX', 'Organization schema', '"mtp" brand 0 CTR', 'P0', 'Phase 0', '/ua/ + /', 'Brand search disambiguation'],
    ['CTR FIX', 'Google Business Profile claim', '"mtp group" queries', 'P0', 'Phase 0', 'GBP external', 'Не сайт — GBP панель'],
]

for r in gaps:
    ws.append(r)

style_header(ws)
freeze(ws, 'A2')
set_widths(ws, [14, 40, 25, 10, 12, 50, 50])

# Priority color coding
for row_idx in range(2, ws.max_row + 1):
    prio = ws[f'D{row_idx}'].value
    if prio == 'P0':
        ws[f'D{row_idx}'].fill = P0_FILL
    elif prio == 'P1':
        ws[f'D{row_idx}'].fill = P1_FILL
    elif prio == 'P2':
        ws[f'D{row_idx}'].fill = P2_FILL

ws.auto_filter.ref = f'A1:G{ws.max_row}'

# ===== SHEET 6: ROADMAP PHASES =====
ws = wb.create_sheet('🗓️ Roadmap Phases')

phases = [
    ['Phase', 'Time', 'Task', 'Owner page / action', 'Priority', 'Status', 'Expected impact', 'Done date', 'Commit', 'Notes'],
    # Phase 0
    ['Phase 0', 'Цей тиждень', 'Переписати home title + meta з ціною', '/ua/ /ru/ /en/', 'P0', 'TODO', '+40-60 clicks/mo'],
    ['Phase 0', 'Цей тиждень', 'Organization schema для brand search', '/ua/ + /', 'P0', 'TODO', '+20-40 clicks/mo'],
    ['Phase 0', 'Цей тиждень', 'Оновити /tsiny/ /tsenu/ /prices/ meta', 'все 3 мови', 'P0', 'TODO', '+15-25 clicks/mo'],
    ['Phase 0', 'Цей тиждень', 'Claim Google Business Profile', 'зовнішнє (не код)', 'P0', 'TODO', 'Brand CTR'],
    # Phase 1
    ['Phase 1', 'Тижні 1-4', 'Pillar /ua/shcho-take-fulfilment/ 3-4K слів', '/ua/shcho-take-fulfilment/', 'P0', 'TODO', '+100-200 clicks/mo через 2-3 міс'],
    ['Phase 1', 'Тижні 1-4', 'Pillar RU версія', '/ru/shcho-take-fulfilment/', 'P0', 'TODO', '+40-80 clicks/mo'],
    ['Phase 1', 'Тижні 1-4', 'Pillar EN версія', '/en/what-is-fulfillment-ukraine/', 'P1', 'TODO', '+20-40 clicks/mo'],
    ['Phase 1', 'Тижні 1-4', 'Blog /scho-take-artykul-sku/', '/ua/blog/scho-take-artykul-sku/', 'P0', 'TODO', '+30-50 clicks/mo'],
    ['Phase 1', 'Тижні 1-4', 'Blog /scho-take-sla-v-logistici/', '/ua/blog/scho-take-sla-v-logistici/', 'P0', 'TODO', '+25-40 clicks/mo'],
    ['Phase 1', 'Тижні 1-4', 'Оптимізувати /ua/fulfilment-ukraina/', 'existing page', 'P0', 'TODO', 'pos 7.3 → 3'],
    ['Phase 1', 'Тижні 1-4', '/ua/fulfilment-dnipro/', '/ua/fulfilment-dnipro/', 'P0', 'TODO', '+15-25 clicks/mo'],
    # Phase 2
    ['Phase 2', 'Тижні 5-8', '/ua/fulfilment-kharkiv/', 'new', 'P1', 'TODO', '+10-20 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', '/ua/fulfilment-lviv/ повна', 'rewrite existing', 'P1', 'TODO', '+15-25 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', '/ua/fulfilment-odesa/ повна', 'rewrite existing', 'P1', 'TODO', '+10-20 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', 'RU blog infrastructure', '/ru/blog/', 'P1', 'TODO', 'Infrastructure only'],
    ['Phase 2', 'Тижні 5-8', 'RU blog: что такое фулфилмент', '/ru/blog/chto-takoye-fulfilment/', 'P1', 'TODO', '+30-50 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', 'RU blog: как открыть магазин', '/ru/blog/kak-otkryt-internet-magazin/', 'P1', 'TODO', '+20-30 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', 'RU blog: SLA (RU)', '/ru/blog/sla-v-logistike/', 'P1', 'TODO', '+15-25 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', 'Marketplace: Rozetka', '/ua/fulfilment-rozetka/', 'P1', 'TODO', '+20-40 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', 'Marketplace: Prom.ua', '/ua/fulfilment-prom/', 'P1', 'TODO', '+15-30 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', 'Marketplace: OLX', '/ua/fulfilment-olx/', 'P1', 'TODO', '+10-20 clicks/mo'],
    ['Phase 2', 'Тижні 5-8', '/ua/3pl-ukraina/', 'new', 'P1', 'TODO', '+15-25 clicks/mo'],
    # Phase 3
    ['Phase 3', 'Місяці 2-3', 'Deep hub: б\'юті (косметика+парфуми+БАДи)', 'consolidation', 'P1', 'TODO', '+30-50 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Deep hub: fashion (одяг+взуття+текстиль)', 'consolidation', 'P1', 'TODO', '+25-40 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Deep hub: electronics', 'consolidation', 'P1', 'TODO', '+20-35 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Deep hub: kids+toys+stationery', 'consolidation', 'P2', 'TODO', '+15-25 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Deep hub: home+furniture+sports', 'consolidation', 'P2', 'TODO', '+15-25 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Deep hub: handmade+gifts+flowers', 'consolidation', 'P2', 'TODO', '+10-20 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Deep hub: auto+construction+heavy', 'consolidation', 'P2', 'TODO', '+10-20 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Deep hub: food+pets+medical', 'consolidation', 'P2', 'TODO', '+10-20 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Публічна /ua/tarify/', 'new', 'P1', 'TODO', '+30-60 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', '"Відповідальне зберігання" hub', '/ua/vidpovidalne-zberigannya/', 'P1', 'TODO', '+40-80 clicks/mo'],
    ['Phase 3', 'Місяці 2-3', 'Amazon FBA Prep EN', '/en/amazon-fba-prep-ukraine/', 'P2', 'TODO', '+10-20 clicks/mo'],
    # Phase 4
    ['Phase 4', 'Місяці 4-6', 'EN backlink campaign', 'external', 'P2', 'TODO', 'topic authority'],
    ['Phase 4', 'Місяці 4-6', 'Case studies EcoDrive + others', '/ua/blog/case-*', 'P2', 'TODO', '+10-20 clicks/mo each'],
    ['Phase 4', 'Місяці 4-6', 'Google Ads + SEO synergy review', 'analytics cross-match', 'P2', 'TODO', 'Paid+organic sync'],
    ['Phase 4', 'Місяці 4-6', 'Content refresh цикл existing pages', 'rolling', 'P2', 'TODO', 'Maintenance'],
    ['Phase 4', 'Місяці 4-6', 'Archive/301 EN pages без трафіку', 'cleanup', 'P2', 'TODO', 'Consolidate signal'],
]

for i, r in enumerate(phases):
    if i == 0:
        ws.append(r)
    else:
        # Lookup progress for this task
        task_name = r[2]
        done = progress_map.get(task_name)
        if done:
            status = '✅ ' + done.get('status', 'DONE')
            r = list(r)
            r[5] = status
            r.extend([done.get('date', ''), done.get('commit', ''), done.get('notes', '')])
        else:
            r = list(r) + ['', '', '']
        ws.append(r)

style_header(ws)
freeze(ws, 'A2')
set_widths(ws, [12, 14, 50, 40, 10, 16, 35, 12, 20, 60])

for row_idx in range(2, ws.max_row + 1):
    phase = ws[f'A{row_idx}'].value
    prio = ws[f'E{row_idx}'].value
    status = ws[f'F{row_idx}'].value or ''
    if phase == 'Phase 0':
        ws[f'A{row_idx}'].fill = P0_FILL
    elif phase == 'Phase 1':
        ws[f'A{row_idx}'].fill = P0_FILL
    elif phase == 'Phase 2':
        ws[f'A{row_idx}'].fill = P1_FILL
    elif phase == 'Phase 3':
        ws[f'A{row_idx}'].fill = P1_FILL
    elif phase == 'Phase 4':
        ws[f'A{row_idx}'].fill = P2_FILL

    if prio == 'P0':
        ws[f'E{row_idx}'].fill = P0_FILL
    elif prio == 'P1':
        ws[f'E{row_idx}'].fill = P1_FILL
    elif prio == 'P2':
        ws[f'E{row_idx}'].fill = P2_FILL

    # Status color
    if 'DONE' in status or 'PARTIAL' in status:
        ws[f'F{row_idx}'].fill = DONE_FILL
        ws[f'F{row_idx}'].font = Font(bold=True, color='1B5E20')
    elif 'MANUAL' in status:
        ws[f'F{row_idx}'].fill = P2_FILL

ws.auto_filter.ref = f'A1:J{ws.max_row}'

# ===== SHEET 7: v1 vs v2 =====
ws = wb.create_sheet('🔄 v1 vs v2')

v1v2 = [
    ['Verdict', 'План v1 (2026-03-27)', 'Реальність / v2 (2026-04-20)', 'Висновок'],
    # Confirmed
    ['✅ CONFIRMED', 'UA-first стратегія', 'UA = 55% всіх кліків (16/29 за 90d)', 'UA-first правильна'],
    ['✅ CONFIRMED', 'Pillar "Що таке фулфілмент" #1 priority', 'Все ще #1 — "фулфілмент" на pos 16', 'Залишається головним таском'],
    ['✅ CONFIRMED', 'Калькулятор важливий', 'Калькулятор ранжується pos 6.1', 'Працює'],
    ['✅ CONFIRMED', '"Для інтернет-магазинів" як buyer intent', '4 кліки за 90d — підтверджено', 'Оптимізувати далі'],
    ['✅ CONFIRMED', 'FAQPage schema для AI visibility', 'FAQ ранжується, дає 1 клік', 'Масштабувати на інші pages'],
    # Refuted
    ['❌ REFUTED', 'EN як пріоритет "🟠 Medium"', '55 EN сторінок → 1 клік за 90d', 'Заморозити EN backlog до deep pillar'],
    ['❌ REFUTED', '22 тонких vertical pages "для X"', '6 зроблених — 0 трафіку', 'Consolidate у 8 deep hubs'],
    ['❌ REFUTED', '"Чому Київ кращий" для Львова/Одеси', 'GSC: реальний search intent на ці міста', 'Повноцінні landings замість відмазки'],
    ['❌ REFUTED', 'Програматичне SEO з 1 шаблону (5-10/day)', 'Thin pages не ранжуються (6 прикладів)', 'Unique deep content > volume'],
    ['❌ REFUTED', '18 EN pages як окремий axis', 'EN vertical strategy провалилась', '1 deep EN pillar замість 18 тонких'],
    # New in v2
    ['🆕 NEW in v2', '—', 'Striking distance оптимізація існуючого (9 opps)', 'v1 не розглядав оптимізацію — тільки нові pages'],
    ['🆕 NEW in v2', '—', 'Zero-CTR paradox як окрема проблема', 'Topic: пошук є, кліків немає — snippet fix'],
    ['🆕 NEW in v2', '—', 'RU blog infrastructure', 'v1 не планував RU blog'],
    ['🆕 NEW in v2', '—', 'Географія Дніпро + Харків', 'v1 мав тільки Київ/Львів/Одеса/Бориспіль'],
    ['🆕 NEW in v2', '—', 'Brand search optimization ("mtp")', 'v1 не вивчав brand signals'],
    ['🆕 NEW in v2', '—', 'Pillar consolidation стратегія', 'v1 йшов на фрагментацію, v2 на consolidation'],
    ['🆕 NEW in v2', '—', 'Marketplace як окрема axis (Rozetka/Prom/OLX)', 'v1 цього не мав'],
    ['🆕 NEW in v2', '—', '"Відповідальне зберігання" content gap', 'Competitor research виявив (500-1K/mo)'],
    ['🆕 NEW in v2', '—', 'Pricing transparency як CTR fix', 'v1 не зв\'язував ціни з CTR'],
]

for r in v1v2:
    ws.append(r)

style_header(ws)
freeze(ws, 'A2')
set_widths(ws, [18, 45, 55, 45])

for row_idx in range(2, ws.max_row + 1):
    verdict = ws[f'A{row_idx}'].value
    if verdict and 'CONFIRMED' in verdict:
        ws[f'A{row_idx}'].fill = DONE_FILL
    elif verdict and 'REFUTED' in verdict:
        ws[f'A{row_idx}'].fill = P0_FILL
    elif verdict and 'NEW' in verdict:
        ws[f'A{row_idx}'].fill = P1_FILL

ws.auto_filter.ref = f'A1:D{ws.max_row}'

# ===== SHEET 8: COMPETITORS =====
ws = wb.create_sheet('🏢 Competitors')

competitors = [
    ['Competitor', 'URL', 'Position in SERP', 'Strengths (semantic axis)', 'Weaknesses', 'MTP action'],
    ['Unipost', 'unipost.ua', 'Top 3 "фулфілмент"', 'Прозорі ціни, калькулятор, marketplace integrations', 'Стара UI, weak blog', 'Beat on blog + CTR'],
    ['Nova Post (NP)', 'novapost.com/uk-ua/for-business/fulfillment/', 'Top 5 brand queries', 'Brand trust, NP ecosystem, courier synergy', 'Thin content, no deep hubs', 'Beat on depth of content'],
    ['Zammler', 'zammler.com.ua', 'Top 10 enterprise queries', 'Enterprise trust, EU focus', 'No SMB language', 'Target SMB/startup axis'],
    ['4erdak', '4erdak.com.ua/service/fulfilment-*', 'Top 5 "малий бізнес"', 'Clear segmentation: для SMB / для інтернет-магазину / відповідальне зберігання', 'Weak brand', 'Replicate audience segmentation'],
    ['PTL Group', 'ptl-group.com.ua/page/fulfilment', 'Top 5-10', 'Clean UX, clear pricing', 'Small catalog', 'Steal UX patterns'],
    ['Diad Logistic', 'diad-logistic.com.ua/sklads-kij-fulfilment', 'Top 10-15', 'Warehouse-first angle', 'Dated content', 'Update warehouse depth'],
    ['Senderukraine', 'senderukraine.com/en/prices', 'EN top-10', 'Transparent pricing EN', 'UA weak', 'Copy pricing transparency'],
    ['Fashion Logistics', 'fashionlogistics.com.ua/en/fulfillment/', 'Vertical niche', 'Fashion-specific', 'Narrow', 'Build fashion deep hub to compete'],
    ['GTAL', 'gtal.com.ua/fulfilment', 'Local niche', 'Boryspil geo-focus', 'Small footprint', 'Use Boryspil airport strength'],
    ['Fulfillment Horoshop', 'horoshop.ua/ua/blog/what-is-fulfillment/', 'Top 3 "що таке фулфілмент"', 'CMS brand authority', 'Not a real fulfilment company', 'Beat with pillar hub + deeper guide'],
    ['KeyCRM Blog', 'blog.keycrm.app/uk/fulfilment-dlya-internet-magazinu-kogo-obrati-v-ukraini/', 'Top 5 listicle', 'Comparison listicles', 'CRM company, not 3PL', 'Create own listicle, be honest'],
    ['Dropplatforma', 'blog.dropplatforma.com.ua/obzory/top-12-fulfilment-kompanij-dlya-internet-magaziniv/', 'Top 10 listicle', 'Affiliate listicle', 'Low quality', 'Outrank with ТОП-X honest review'],
    ['Cpashka', 'cpashka.biz/blog/strong-top-11-fulfilment-operatoriv-ukrainy-strong/', 'Top 10 listicle', 'Listicle volume', 'Low authority', 'Outrank with ТОП-X'],
    ['KeepinCRM', 'keepincrm.com/fullfillment-companies-in-ukraine', 'Top 10', 'Comparison table', 'CRM not 3PL', 'Position MTP в їхній таблиці чи зробити свою'],
    ['Rozetka', 'rozetka.com.ua (news: запустили FBO)', 'Top 1-2 for "rozetka fulfillment"', 'Marketplace itself launched FBO', 'Not open 3PL', 'Position as "alternative to Rozetka FBO"'],
    ['', '', '', '', '', ''],
    ['BOTTOM LINE', '', '', '', '', ''],
    ['MTP не в топ-10 за "фулфілмент для інтернет-магазину"', '', '', '', '', 'Top priority — beat pillar'],
    ['3 pillar gaps MTP:', '', '', '', '', ''],
    ['A: educational "що таке фулфілмент" — 1-2K/mo', '', '', '', '', 'Phase 1 pillar'],
    ['B: "ТОП операторів" comparison listicles', '', '', '', '', 'Phase 3'],
    ['C: marketplace integrations (Rozetka/Prom/OLX)', '', '', '', '', 'Phase 2'],
    ['', '', '', '', '', ''],
    ['TOTAL GAP: ~5,000-8,000 monthly UA queries де MTP не в топ-10', '', '', '', '', ''],
]

for r in competitors:
    ws.append(r)

style_header(ws)
freeze(ws, 'A2')
set_widths(ws, [22, 50, 20, 45, 35, 45])

# Bold "BOTTOM LINE" section
for row_idx in range(17, ws.max_row + 1):
    ws[f'A{row_idx}'].font = Font(bold=True)

ws.auto_filter.ref = f'A1:F15'  # only competitor table

# ===== SHEET 9: TOP QUERIES =====
ws = wb.create_sheet('🔍 Top queries')

queries_file = os.path.join(ROOT, 'docs', 'gsc', 'full-queries.csv')
queries = []
with open(queries_file, encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for r in reader:
        queries.append(r)

ws.append(['Query', 'Clicks 90d', 'Impressions 90d', 'CTR %', 'Avg position', 'Intent classification', 'Priority'])

def classify_intent(q):
    q = q.lower()
    if any(w in q for w in ['ціна', 'цена', 'вартість', 'price', 'cost', 'тариф', 'скільки коштує', 'сколько стоит']):
        return 'COMMERCIAL (price)'
    if any(w in q for w in ['mtp', 'мтп', 'фулфілмент мтп']):
        return 'BRAND'
    if any(w in q for w in ['що таке', 'что такое', 'what is']):
        return 'INFORMATIONAL (edu)'
    if any(w in q for w in ['як', 'как', 'how to', 'how']):
        return 'INFORMATIONAL (howto)'
    if any(w in q for w in ['vs', 'проти', 'против', 'або', 'или', 'топ', 'top', 'кращий', 'лучший', 'best', 'порівняння']):
        return 'COMPARATIVE'
    if any(w in q for w in ['фулф', 'fulfill', '3pl', 'логістик', 'логистик', 'logistic', 'warehouse', 'склад']):
        return 'COMMERCIAL (buyer)'
    if any(w in q for w in ['rozetka', 'prom', 'olx', 'kasta', 'shopify', 'amazon', 'etsy', 'rozetka']):
        return 'CHANNEL/PLATFORM'
    if any(w in q for w in ['київ', 'киев', 'kyiv', 'kiev', 'львів', 'львов', 'lviv', 'одеса', 'одесса', 'odesa', 'харків', 'харьков', 'kharkiv', 'дніпро', 'днепр', 'dnipro', 'бориспіль', 'борисполь', 'boryspil']):
        return 'GEO'
    return 'OTHER'

def priority_query(row):
    try:
        clk = int(row['clicks_90d'])
        imp = int(row['impressions_90d'])
        pos = float(row['avg_position']) if row['avg_position'] else 99
    except:
        clk, imp, pos = 0, 0, 99

    if imp >= 100 and 4 <= pos <= 20:
        return 'P0'
    if imp >= 50 and pos > 20:
        return 'P1'
    if clk >= 2:
        return 'P0'
    if imp >= 30:
        return 'P1'
    return 'P2'

for r in queries:
    q = r['query']
    ws.append([
        q,
        int(r['clicks_90d']) if r['clicks_90d'] else 0,
        int(r['impressions_90d']) if r['impressions_90d'] else 0,
        float(r['ctr_pct']) if r['ctr_pct'] else 0,
        float(r['avg_position']) if r['avg_position'] else 99,
        classify_intent(q),
        priority_query(r)
    ])

style_header(ws)
freeze(ws, 'A2')
set_widths(ws, [40, 12, 16, 10, 12, 25, 10])

for row_idx in range(2, ws.max_row + 1):
    prio = ws[f'G{row_idx}'].value
    if prio == 'P0':
        ws[f'G{row_idx}'].fill = P0_FILL
    elif prio == 'P1':
        ws[f'G{row_idx}'].fill = P1_FILL
    elif prio == 'P2':
        ws[f'G{row_idx}'].fill = P2_FILL

ws.auto_filter.ref = f'A1:G{ws.max_row}'

# ===== SHEET 10: COMPLETION LOG =====
ws = wb.create_sheet('📅 Completion log')
ws['A1'] = 'COMPLETION LOG — хронологія виконаних задач'
ws['A1'].font = Font(bold=True, size=14, color='E63329')
ws.merge_cells('A1:F1')

ws.append([])
ws.append(['Date', 'Task', 'Status', 'Commit', 'Phase', 'Notes'])
style_header(ws, row=3)

# Sort completions by date desc (newest first)
sorted_done = sorted(progress_map.values(), key=lambda c: c.get('date') or '0', reverse=True)

for c in sorted_done:
    status = c.get('status', 'DONE')
    # Try to infer phase from task name (best-effort)
    task = c['task']
    phase = ''
    if any(k in task.lower() for k in ['home title', 'organization schema', 'tsiny', 'gbp', 'google business']):
        phase = 'Phase 0'
    elif any(k in task.lower() for k in ['generate_lead', 'noscript', 'tilda']):
        phase = 'Maintenance'
    elif 'gsc' in task.lower() or 'competitor' in task.lower():
        phase = 'Analysis'

    ws.append([
        c.get('date', '—'),
        task,
        '✅ ' + status,
        c.get('commit', ''),
        phase,
        c.get('notes', '')[:200]
    ])

freeze(ws, 'A4')
set_widths(ws, [12, 55, 22, 20, 14, 80])

for row_idx in range(4, ws.max_row + 1):
    status = ws[f'C{row_idx}'].value or ''
    if 'DONE' in status:
        ws[f'C{row_idx}'].fill = DONE_FILL
    elif 'PARTIAL' in status:
        ws[f'C{row_idx}'].fill = P1_FILL
    elif 'MANUAL' in status:
        ws[f'C{row_idx}'].fill = P2_FILL

ws.auto_filter.ref = f'A3:F{ws.max_row}'

# ===== SAVE =====
wb.save(OUT)
print(f'Saved: {OUT}')
print(f'Sheets: {wb.sheetnames}')
print(f'Completions tracked: {len(progress_map)}')
print(f'Size: {os.path.getsize(OUT)/1024:.1f} KB')
